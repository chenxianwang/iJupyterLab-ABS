# -*- coding: utf-8 -*-
'''
Created on Sun Jun  4 17:31:54 2017

@author: Jonah.Chen
'''
import os.path
import datetime
import pandas as pd
import numpy as np
from openpyxl import Workbook,load_workbook
from dateutil.relativedelta import relativedelta
from datetime import date
import logging
import sqlalchemy
from constant import * 

pd.options.mode.chained_assignment = None
    
def get_logger(logger_name):
    
    logging.basicConfig(#filename=path_root + '/asset_pool_analysis.txt',\
                        level = logging.DEBUG,\
                        #format = '%(asctime)s - %(name)s - %(levelname)s - %(message)s'
                        )
    return logging.getLogger(logger_name)

def save_to_excel(df,ws_name,wb_name):
    
    get_logger(__name__).info('Saving {0}'.format(ws_name))
    
    if not os.path.isfile(wb_name):
            Workbook().save(wb_name)
        
    writer = pd.ExcelWriter(wb_name, engine = 'openpyxl')
    writer.book = load_workbook(wb_name)
    
    if ws_name not in [ws.title for ws in writer.book.worksheets]:
        writer.book.create_sheet(ws_name)
    
    try:
        writer.book.remove_sheet(writer.book.get_sheet_by_name('Sheet'))
    except(KeyError):
        pass
    
    writer.sheets = dict((ws.title,ws) for ws in writer.book.worksheets)

    if not isinstance(df,list):
        df = [df]
    
    for df_content in df:
        rows_used = writer.book.get_sheet_by_name(ws_name).max_row
        if rows_used == 1:                                   
            df_content.to_excel(writer,sheet_name = ws_name,index=False)
        else:
            df_content.to_excel(writer,sheet_name = ws_name,startrow = rows_used + 2,index=False) 

    writer.save()

def save_to_mysql(df,schema_name,table_name,action_if_exists):
    
    engine = sqlalchemy.create_engine('mysql+mysqldb://root:hccfc_abs@localhost/'+ schema_name + '?charset=utf8', 
                                      convert_unicode=True)
    df.to_sql(con=engine, name = table_name,if_exists = action_if_exists,index=False) # append or replace

def percentage_format(df):
    
    for elem in list(df):
        df[elem] = pd.Series(['{0:.2f}%'.format(val * 100) for val in df[elem]], index = df.index)
    return df


def decimal_format(df):
    for elem in list(df):
        df[elem] = pd.Series([round(val, 2) for val in df[elem]], index = df.index)
    return df

def get_next_eom(dt,month_increment):
    
    month_day_nonleap = [31,28,31,30,31,30,31,31,30,31,30,31]
    month_day_leap = [31,29,31,30,31,30,31,31,30,31,30,31]

    dt_next_month = dt + relativedelta(months=month_increment)
    
    if dt_next_month.year % 4 == 0:
        month_day = month_day_leap
    else : month_day = month_day_nonleap
    
    dt_next_eom = date(dt_next_month.year,dt_next_month.month,month_day[dt_next_month.month - 1])
    
    return dt_next_eom

def get_next_eoq(dt):
    Q = (dt.month-1)//3 + 1
    candidates = [
        datetime.date(dt.year - 1, 12, 31),
        datetime.date(dt.year, 3, 31),
        datetime.date(dt.year, 6, 30),
        datetime.date(dt.year, 9, 30),
        datetime.date(dt.year, 12, 31),
    ]
    return candidates[Q-1]

def SD_with_weight(value_list,weight_list):
    
    value_list = np.array(value_list)
    weight_list = np.array(weight_list)
    mean_value = np.multiply(value_list,weight_list).sum()
    SD = np.sqrt(np.multiply( (value_list - mean_value)**2 , weight_list).sum())
    return SD

def drop_outliers(df,indicators,n_sigma):
    for indicator in indicators:
        df = df[np.abs(df[indicator]-df[indicator].mean())<=(n_sigma*df[indicator].std())]
    return df


def df_bins_result(df,d,bins_this):
    d_cut = pd.cut(df[d], bins=bins_this)
    return stastics_group_by_d(df,d_cut,d)

def stastics_group_by_d(df,d_cut,d):
    df_this = df.groupby([d_cut])\
               .agg({'Amount_Contract':'sum',
                     'Amount_Outstanding':'sum','No_Contract':'count'})\
               .reset_index()\
               .rename(columns = {'Amount_Contract':'合同本金（万元）',
                                  'Amount_Outstanding':'本金余额（万元）',
                                  'No_Contract':'贷款笔数'                                                      
                                  }
                       )
    df_this = percentage_cal(df_this,d)
    return df_this
  
def percentage_cal(df,d):
    
    df[['合同本金（万元）','本金余额（万元）']] = df[['合同本金（万元）','本金余额（万元）']].where(df['贷款笔数']>0,0)
    
    df['合同本金占比'] = df['合同本金（万元）'] / sum(df['合同本金（万元）'])
    df['本金余额占比'] = df['本金余额（万元）'] / sum(df['本金余额（万元）'])
    df['贷款笔数占比'] = df['贷款笔数'] / sum(df['贷款笔数'])
    
    df.loc['总计']= df.sum()
    df.iloc[-1,df.columns.get_loc(d)] = '总计'
    df['平均每笔余额（元）'] = 10000 * df['本金余额（万元）'] / df['贷款笔数']
    
    df[['合同本金占比','本金余额占比','贷款笔数占比']] = percentage_format(df[['合同本金占比','本金余额占比','贷款笔数占比']])
    df[['合同本金（万元）','本金余额（万元）','平均每笔余额（元）']] = decimal_format(df[['合同本金（万元）','本金余额（万元）','平均每笔余额（元）']])
    
    return df

def Condition_Satisfied_or_Not(Assets,target_d,Targets):
        if target_d in ['Credit_Score_max','LoanRemainTerm','Interest_Rate_max']:
            value_d = sum(Assets[target_d] * Assets['Amount_Outstanding']) / sum(Assets['Amount_Outstanding'])
            if  value_d <= Targets[target_d]['object_value'] :
                print(target_d + ' Condition Completed. Weighted Average of ' + target_d + ' update:',value_d)
                print('Target is...',Targets[target_d]['object'],Targets[target_d]['object_value'])
            else:            
                print(target_d + ' Condition NOT Completed. Weighted Average of ' + target_d + ' update:',value_d)
                print('Target is...',Targets[target_d]['object'],Targets[target_d]['object_value'])
        elif target_d in ['SCp']:
            value_d = sum(Assets[Assets[target_d] ==1]['Amount_Outstanding']) / sum(Assets['Amount_Outstanding'])
            if  value_d >= Targets[target_d]['object_value'] :
                print(target_d + ' Condition Completed. ' + target_d + ' update:',value_d)
                print('Target is...',Targets[target_d]['object'],Targets[target_d]['object_value'])
            else:            
                print(target_d + ' Condition NOT Completed. ' + target_d + ' update:',value_d)
                print('Target is...',Targets[target_d]['object'],Targets[target_d]['object_value'])
                
        elif target_d in ['Credit_Score_min','Interest_Rate_min']:
            value_d = sum(Assets[target_d] * Assets['Amount_Outstanding']) / sum(Assets['Amount_Outstanding'])
            if  value_d >= Targets[target_d]['object_value'] :
                print(target_d + ' Condition Completed. ' + target_d + ' update:',value_d)
                print('Target is...',Targets[target_d]['object'],Targets[target_d]['object_value'])
            else:            
                print(target_d + ' Condition NOT Completed. ' + target_d + ' update:',value_d)
                print('Target is...',Targets[target_d]['object'],Targets[target_d]['object_value'])
                
        else: print('OutstandingPrincipal Condition Completed. OutstandingPrincipal update:',sum(Assets['Amount_Outstanding']))
#            if (sum(Assets['Amount_Outstanding']) <= Targets['Amount_Outstanding_max']['object_value']) & (sum(Assets['Amount_Outstanding']) >= Targets['Amount_Outstanding_min']['object_value']):
#                print('OutstandingPrincipal Condition Completed. OutstandingPrincipal update:',sum(Assets['Amount_Outstanding']))
#                print('Target is...',Targets['Amount_Outstanding_max']['object'],Targets['Amount_Outstanding_max']['object_value'], ' and ', Targets['Amount_Outstanding_min']['object'],Targets['Amount_Outstanding_min']['object_value'] )
#            else:            
#                print('!!! OutstandingPrincipal Condition NOT Completed !!! OutstandingPrincipal update:',sum(Assets['Amount_Outstanding']))
#                print('Target is...',Targets['Amount_Outstanding_max']['object'],Targets['Amount_Outstanding_max']['object_value'], ' and ', Targets['Amount_Outstanding_min']['object'],Targets['Amount_Outstanding_min']['object_value'] )


